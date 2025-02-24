from openpyxl import load_workbook

class ExcelService:
    header_row = 1
    def __init__(self, file_name: str, sheet: str):
        self.file_name = file_name
        self.sheet = sheet

    def _load(self):
        self.wb = load_workbook(self.file_name, data_only=True)
        self.ws = self.wb[self.sheet]

    def _append_result_columns(self):
        self.write_data(row=self.header_row, column=self.ws.max_column + 1, value="Tên ngân hàng")
        self.write_data(row=self.header_row, column=self.ws.max_column + 1, value="Kết quả")

    def prepare_wb(self):
        self._load()
        self._append_result_columns()

    def write_data(self, row: int, column: int, value:str):
        self.ws.cell(row=row, column=column, value=value)

    def read_data(self, row: int, column: int) -> str:

        return self.ws.cell(row=row, column=column).value
    
    def save(self):
        self.wb.save(self.file_name)

    @property
    def get_max_row(self) -> int:
        return self.ws.max_row
    @property
    def get_max_column(self) -> int:
        return self.ws.max_column
    